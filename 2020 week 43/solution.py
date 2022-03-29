from openpyxl import load_workbook
import pandas as pd
from pandasgui import show


def get_names(workbook, gender):
    df = None
    for entry, data_boundary in workbook.tables.items():
        # parse the data within the ref boundary
        data = workbook[data_boundary]
        # extract the data
        # the inner list comprehension gets the values for each cell in the table
        content = [[cell.value for cell in ent]for ent in data]

        header = content[0]

        # the contents ... excluding the header
        rest = content[1:]

        # create dataframe with the column names
        # and pair table name with dataframe
        df = pd.concat([df ,pd.DataFrame(rest, columns=header).assign(Month=entry, Gender=gender)])
    return df

df_girls = get_names(load_workbook('2019girlsnames.xlsx')['Table 5'], 'Girls')
df_boys =  get_names(load_workbook('2019boysnames.xlsx')['Table 5'], 'Boys')
df_output_1 = pd.concat([df_boys, df_girls])

df_output_2 = df_output_1.groupby('Name', as_index=False)['Count'].sum().assign(Rank = lambda dfx : dfx['Count'].rank(ascending=False))
show(df_output_2)